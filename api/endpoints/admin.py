"""Admin API endpoints"""
import traceback
import uuid
import re
import io
import time
import threading
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from app.auth.auth_dependency import verify_manager_api_key as verify_firebase_token
from db.session import get_db_connection, close_db_connection
from crud import admin as admin_crud
from crud import store as store_crud
from crud import menu as menu_crud
from crud import terms as terms_crud
from models.store import StoreCreate
from models.menu import Menu
from core.s3_config import S3_CLIENT, BUCKET_NAME, TERMS_BUCKET_NAME
import os
import openpyxl
from copy import copy

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '../../templates/contract_template.xlsx')

router = APIRouter()


def _normalize_phone(phone: str) -> str:
    """전화번호를 010-XXXX-XXXX 형식으로 정규화"""
    if not phone:
        return ''
    digits = re.sub(r'\D', '', phone)
    if digits.startswith('82'):
        digits = '0' + digits[2:]
    if not digits.startswith('010'):
        return phone
    if len(digits) == 11:
        return f'{digits[:3]}-{digits[3:7]}-{digits[7:]}'
    return phone


_dashboard_cache: dict = {"data": None, "expires_at": 0}
_dashboard_cache_lock = threading.Lock()
_DASHBOARD_CACHE_TTL = 30  # seconds

@router.get("/dashboard/statistics")
def get_dashboard_statistics(user=Depends(verify_firebase_token)):
    """대시보드 통계 데이터 (30초 캐싱)"""
    now = time.time()
    with _dashboard_cache_lock:
        if _dashboard_cache["data"] is not None and now < _dashboard_cache["expires_at"]:
            return _dashboard_cache["data"]

    connection = get_db_connection()
    try:
        result = admin_crud.get_dashboard_statistics(connection)
        with _dashboard_cache_lock:
            _dashboard_cache["data"] = result
            _dashboard_cache["expires_at"] = time.time() + _DASHBOARD_CACHE_TTL
        return result
    except Exception as e:
        print(f"Error in get_dashboard_statistics: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


# ── 플랫폼 대시보드 API (GNB-164 / GNB-165 / GNB-166) ───────────────────────

@router.get("/dashboard/summary")
def get_dashboard_summary(user=Depends(verify_firebase_token)):
    """실시간 요약: 발행잔액 / 이번 정산주기 예정 / 누적 지표

    - issued_balance: 미사용 기프티콘 menu.price 합계 (REFUNDED/CANCELED 제외)
    - current_cycle: settlement_details 미연결 건 기준 실시간 예상값
    - cumulative: stats_daily_platform 전체 SUM (매일 00:10 배치 갱신)
    """
    try:
        from crud import stats as stats_crud
        return stats_crud.get_dashboard_summary()
    except Exception as e:
        print(f"Error in get_dashboard_summary: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/dashboard/stats")
def get_dashboard_stats(
    period: str = Query('monthly', description="집계 단위: daily | weekly | monthly | yearly | all"),
    page: int = Query(1, ge=1, description="페이지 번호 (daily/weekly/monthly에서만 적용)"),
    size: int = Query(30, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """기간별 운영 통계 (stats_daily_platform 기반)

    GNB-169:
    - daily/weekly/monthly: 최신순 30개 페이지네이션, total_row(전체 합계) 항상 포함
    - yearly/all: 페이지네이션 없음
    - 수수료: PG 수수료 차감 후 순수수료 (배치 집계값)
    - weekly 라벨: 2026-01-01~2026-01-07 형식
    """
    try:
        from crud import stats as stats_crud
        return stats_crud.get_dashboard_stats(period, page=page, size=size)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error in get_dashboard_stats: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/dashboard/settlement-cycles")
def get_dashboard_settlement_cycles(
    page: int = Query(1, ge=1, description="페이지 번호"),
    size: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """정산 주기별 플랫폼 매출 이력

    - settlement 테이블 GROUP BY cycle_id로 집계 (별도 요약 테이블 없음)
    - total_settlement_amount: COMPLETED/PENDING 매장 net_payout 합계
    - platform_fee_amount/vat: original_fee 기준 (프로모션 적용 전 수수료)
    - unused_amount: 해당 주기 발행 기프티콘 중 미사용 상태 menu.price 합계
    """
    try:
        from crud import stats as stats_crud
        return stats_crud.get_dashboard_settlement_cycles(page, size)
    except Exception as e:
        print(f"Error in get_dashboard_settlement_cycles: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stores/export/excel")
def export_stores_excel(
    search: Optional[str] = Query(None),
    inspection_status: Optional[str] = Query(None),
    contract_completed: Optional[str] = Query(None),
    user=Depends(verify_firebase_token),
):
    """계약서 발송용 엑셀 파일 생성 - 양식 템플릿 기반"""
    connection = get_db_connection()
    try:
        import pymysql
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        query = '''
            SELECT
                s.store_name,
                s.store_address,
                s.business_number,
                o.name AS owner_name,
                o.email AS owner_email,
                o.phone AS owner_phone
            FROM store s
            LEFT JOIN owner o ON s.owner_id = o.id
        '''
        conditions = []
        params = []

        if search:
            conditions.append('(s.store_name LIKE %s OR o.name LIKE %s)')
            pattern = f'%{search}%'
            params += [pattern, pattern]
        if inspection_status:
            conditions.append('s.inspection_status = %s')
            params.append(inspection_status)
        if contract_completed:
            conditions.append('s.contract_completed = %s')
            params.append(contract_completed)

        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        query += ' ORDER BY s.created_at DESC'

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # 양식 템플릿을 복사해서 데이터 채우기
        template_path = os.path.abspath(TEMPLATE_PATH)
        wb = openpyxl.load_workbook(template_path)
        ws = wb['대량발송']

        # 데이터는 12행부터 입력 (1~11행은 유의사항+헤더, 수정 금지)
        DATA_START_ROW = 12
        for row_idx, row in enumerate(rows, start=DATA_START_ROW):
            # A: 수신자, B: 이메일, C: 휴대폰번호, D: 비밀번호(공란)
            # E: 폼 이름(공란), F: 주소, G: 상호, H: 이름, I: 사업자번호
            ws.cell(row=row_idx, column=1, value=row.get('owner_name') or '')
            ws.cell(row=row_idx, column=2, value=row.get('owner_email') or '')
            ws.cell(row=row_idx, column=3, value=_normalize_phone(row.get('owner_phone') or ''))
            ws.cell(row=row_idx, column=4, value='')
            ws.cell(row=row_idx, column=5, value='')
            ws.cell(row=row_idx, column=6, value=row.get('store_address') or '')
            ws.cell(row=row_idx, column=7, value=row.get('store_name') or '')
            ws.cell(row=row_idx, column=8, value=row.get('owner_name') or '')
            ws.cell(row=row_idx, column=9, value=row.get('business_number') or '')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from urllib.parse import quote
        filename = '기프넛_계약서_발송목록.xlsx'
        encoded_filename = quote(filename, safe='')
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f"attachment; filename=\"contract.xlsx\"; filename*=UTF-8''{encoded_filename}"
            }
        )
    except Exception as e:
        print(f"Error in export_stores_excel: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/stores")
def get_stores(
    search: Optional[str] = Query(None, description="매장 이름, 사장님 이름으로 검색"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    inspection_status: Optional[str] = Query(None, description="승인상태 필터: PENDING, APPROVED, REJECTED, CLOSED"),
    contract_completed: Optional[str] = Query(None, description="계약상태 필터: NONE, SENT, COMPLETED"),
    user=Depends(verify_firebase_token),
):
    """매장 리스트 (관리자용, 페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_stores(connection, search, page, limit, inspection_status, contract_completed)
        return {
            'stores': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_stores: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/stores/{store_id}")
def get_store_detail(store_id: int, user=Depends(verify_firebase_token)):
    """매장 상세 정보"""
    connection = get_db_connection()
    try:
        store = admin_crud.get_store_detail(connection, store_id)
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        return store
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_store_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/stores/{store_id}/menu")
def get_store_menus(
    store_id: int,
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(10, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """매장 메뉴 리스트 (페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_store_menus(connection, store_id, page, limit)
        return {
            'menus': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages'],
            }
        }
    except Exception as e:
        print(f"Error in get_store_menus: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/stores/{store_id}/giftcards")
def get_store_giftcards(
    store_id: int,
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(10, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """매장의 깊티(기프티콘) 리스트 (페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_store_giftcards(connection, store_id, page, limit)
        return {
            'giftcards': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_store_giftcards: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/users")
def get_users(
    search: Optional[str] = Query(None, description="이름, 아이디, 전화번호, ID로 검색"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """유저 리스트 (관리자용, 페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_users(connection, search, page, limit)
        return {
            'users': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_users: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/users/{user_id}")
def get_user_detail(user_id: int, user=Depends(verify_firebase_token)):
    """유저 상세 정보"""
    connection = get_db_connection()
    try:
        user = admin_crud.get_user_detail(connection, user_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        return user
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_user_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/users/{user_id}/orders")
def get_user_orders(user_id: int, user=Depends(verify_firebase_token)):
    """유저 주문 내역"""
    connection = get_db_connection()
    try:
        orders = admin_crud.get_user_orders(connection, user_id)
        return {'orders': orders}
    except Exception as e:
        print(f"Error in get_user_orders: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/users/{user_id}/giftcards")
def get_user_giftcards(user_id: int, user=Depends(verify_firebase_token)):
    """유저 기프티콘 리스트"""
    connection = get_db_connection()
    try:
        giftcards = admin_crud.get_user_giftcards(connection, user_id)
        return {'giftcards': giftcards}
    except Exception as e:
        print(f"Error in get_user_giftcards: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/orders")
def get_orders(
    search: Optional[str] = Query(None, description="주문번호, user id로 검색"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """주문 리스트 (관리자용, 페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_orders(connection, search, page, limit)
        return {
            'orders': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_orders: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/orders/{order_id}")
def get_order_detail(order_id: int, user=Depends(verify_firebase_token)):
    """주문 상세 정보"""
    connection = get_db_connection()
    try:
        order = admin_crud.get_order_detail(connection, order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        return order
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_order_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/orders/{order_id}/giftcards")
def get_order_giftcards(order_id: int, user=Depends(verify_firebase_token)):
    """주문의 기프티콘 리스트"""
    connection = get_db_connection()
    try:
        giftcards = admin_crud.get_order_giftcards(connection, order_id)
        return {'giftcards': giftcards}
    except Exception as e:
        print(f"Error in get_order_giftcards: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/menus")
def get_all_menus(
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """전체 메뉴 리스트 (페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_all_menus(connection, page, limit)
        return {
            'menus': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_all_menus: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/notices")
def get_notices(
    target: Optional[str] = Query(None, description="'user' 또는 'owner', None이면 둘 다"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """공지사항 리스트 (페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_notices(connection, target, page, limit)
        return {
            'notices': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages']
            }
        }
    except Exception as e:
        print(f"Error in get_notices: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.post("/notices")
def create_notice(notice: dict, user=Depends(verify_firebase_token)):
    """공지사항 등록
    
    Body:
        - target: 'user' 또는 'owner'
        - title: 공지사항 제목
        - content: 공지사항 내용
    """
    connection = get_db_connection()
    try:
        target = notice.get('target')
        title = notice.get('title')
        content = notice.get('content')
        
        if not target or target not in ['user', 'owner']:
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        
        if not title or not content:
            raise HTTPException(status_code=400, detail="title and content are required")
        
        notice_id = admin_crud.create_notice(connection, target, title, content)
        return {'message': 'Notice created successfully', 'id': notice_id}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in create_notice: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/notices/{target}/{notice_id}")
def get_notice_detail(target: str, notice_id: int, user=Depends(verify_firebase_token)):
    """공지사항 상세 조회"""
    connection = get_db_connection()
    try:
        if target not in ['user', 'owner']:
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        
        notice = admin_crud.get_notice_detail(connection, target, notice_id)
        if not notice:
            raise HTTPException(status_code=404, detail="Notice not found")
        return notice
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_notice_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.put("/notices/{target}/{notice_id}")
def update_notice(target: str, notice_id: int, notice: dict, user=Depends(verify_firebase_token)):
    """공지사항 수정"""
    connection = get_db_connection()
    try:
        if target not in ['user', 'owner']:
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        
        title = notice.get('title')
        content = notice.get('content')
        
        success = admin_crud.update_notice(connection, target, notice_id, title, content)
        if not success:
            raise HTTPException(status_code=404, detail="Notice not found")
        return {'message': 'Notice updated successfully'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_notice: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.delete("/notices/{target}/{notice_id}")
def delete_notice(target: str, notice_id: int, user=Depends(verify_firebase_token)):
    """공지사항 삭제"""
    connection = get_db_connection()
    try:
        if target not in ['user', 'owner']:
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        
        success = admin_crud.delete_notice(connection, target, notice_id)
        if not success:
            raise HTTPException(status_code=404, detail="Notice not found")
        return {'message': 'Notice deleted successfully'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in delete_notice: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


# ---------- 약관 관리 ----------

@router.post("/terms")
def create_term(body: dict, user=Depends(verify_firebase_token)):
    """약관 종류 추가. Body: target('user'|'owner'), term_type, title, required(optional, default True)"""
    connection = get_db_connection()
    try:
        target = body.get("target", "user")
        term_type = body.get("term_type")
        title = body.get("title")
        required = body.get("required", True)
        if not term_type or not title:
            raise HTTPException(status_code=400, detail="term_type and title are required")
        if target not in ("user", "owner"):
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        term_id = terms_crud.create_term(connection, term_type, title, required, target)
        return {"message": "Term created", "term_id": term_id}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in create_term: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.post("/terms/{term_id}/version")
def create_term_version(term_id: int, body: dict, user=Depends(verify_firebase_token)):
    """약관 버전 추가. 이미 같은 버전이 있으면 기존 version_id로 성공 반환(재등록/재업로드 허용)."""
    connection = get_db_connection()
    try:
        version = body.get("version")
        notice_date = body.get("notice_date")
        effective_date = body.get("effective_date")
        reagreement_required = body.get("reagreement_required", True)
        if not version or not notice_date or not effective_date:
            raise HTTPException(status_code=400, detail="version, notice_date, effective_date are required")
        if isinstance(notice_date, str):
            from datetime import datetime
            notice_date = datetime.strptime(notice_date, "%Y-%m-%d").date()
        if isinstance(effective_date, str):
            from datetime import datetime
            effective_date = datetime.strptime(effective_date, "%Y-%m-%d").date()
        try:
            version_id = terms_crud.create_term_version(
                connection, term_id, version, notice_date, effective_date, reagreement_required
            )
            return {"message": "Term version created", "term_version_id": version_id}
        except ValueError as e:
            if "이미 존재하는 버전입니다" not in str(e):
                raise HTTPException(status_code=400, detail=str(e))
            existing_id = terms_crud.get_term_version_id_by_version(connection, term_id, version)
            if existing_id is None:
                raise HTTPException(status_code=500, detail="Version exists but could not be retrieved")
            return {"message": "Term version already exists", "term_version_id": existing_id}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in create_term_version: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/terms/all")
def get_all_terms(target: Optional[str] = Query(None, description="'user' 또는 'owner', None이면 전체"), user=Depends(verify_firebase_token)):
    """모든 약관 종류 및 버전 조회 (관리자용). target 지정 시 해당만."""
    connection = get_db_connection()
    try:
        if target and target not in ("user", "owner"):
            raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
        result = terms_crud.get_all_terms_with_versions(connection, target)
        return {"terms": result}
    except Exception as e:
        print(f"Error in get_all_terms: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.put("/terms_version/{version_id}")
def update_term_version(version_id: int, body: dict, user=Depends(verify_firebase_token)):
    """약관 버전 수정. Body: optional version, notice_date, effective_date, reagreement_required. 시행일은 공지일+30일 유지."""
    connection = get_db_connection()
    try:
        notice_date = body.get("notice_date")
        effective_date = body.get("effective_date")
        if notice_date and isinstance(notice_date, str):
            from datetime import datetime
            notice_date = datetime.strptime(notice_date, "%Y-%m-%d").date()
        if effective_date and isinstance(effective_date, str):
            from datetime import datetime
            effective_date = datetime.strptime(effective_date, "%Y-%m-%d").date()
        success = terms_crud.update_term_version(
            connection, version_id,
            version=body.get("version"),
            notice_date=notice_date,
            effective_date=effective_date,
            reagreement_required=body.get("reagreement_required"),
        )
        if not success:
            raise HTTPException(status_code=404, detail="Term version not found")
        return {"message": "Term version updated"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_term_version: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/terms/content")
def get_term_content(
    target: str = Query(..., description="user 또는 owner"),
    filename: str = Query(..., description="파일명 예: service_term_260101.html"),
    user=Depends(verify_firebase_token),
):
    """S3에서 약관 본문 파일 조회. key: terms/{user|partner}/{filename}"""
    if target not in ("user", "owner"):
        raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="invalid filename")
    parsed = _parse_terms_filename(filename, target)
    if not parsed:
        raise HTTPException(status_code=400, detail="invalid filename format")
    key = _build_s3_key(target, parsed["term_type"], filename)
    bucket = TERMS_BUCKET_NAME
    try:
        obj = S3_CLIENT.get_object(Bucket=bucket, Key=key)
        body = obj["Body"].read()
        content = body.decode("utf-8", errors="replace")
        return {"content": content, "filename": filename}
    except Exception as e:
        err_str = str(e)
        err_lower = err_str.lower()
        # 조회 실패 시 사용한 bucket/key 로그 (S3 콘솔에서 경로 확인용)
        print(f"[terms/content] S3 GET failed: bucket={bucket}, key={key}, error={err_str}")
        if "nosuchkey" in err_lower or "404" in err_lower or "no such key" in err_lower:
            raise HTTPException(
                status_code=404,
                detail=f"Term file not found. Check S3 key: s3://{bucket}/{key}",
            )
        if "accessdenied" in err_lower or "forbidden" in err_lower:
            raise HTTPException(status_code=403, detail="S3 access denied")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=err_str)


# 파일명 prefix -> term_type
_FILENAME_PREFIX_TO_TERM_TYPE_USER = {
    "service_term": "SERVICE",
    "marketing_term": "MARKETING",
    "privacy_term": "PRIVACY",
    "privacy_consent_term": "PRIVACY_CONSENT",
    "location_term": "LOCATION",
}
_FILENAME_PREFIX_TO_TERM_TYPE_PARTNER = {
    "partner_service_term": "SERVICE",
    "partner_marketing_term": "MARKETING",
    "partner_privacy_term": "PRIVACY",
    "partner_privacy_consent_term": "PRIVACY_CONSENT",
    "partner_location_term": "LOCATION",
    "partner_fee_term": "FEE",
}
_TERM_TYPE_DEFAULT_TITLE = {
    "SERVICE": "서비스 이용약관",
    "MARKETING": "마케팅 정보 수신 이용약관",
    "PRIVACY": "개인정보 처리방침",
    "PRIVACY_CONSENT": "개인정보 수집 동의",
    "LOCATION": "위치정보 이용약관",
    "FEE": "수수료 약관",
}
# term_type -> S3 카테고리 폴더명
_TERM_TYPE_TO_CATEGORY = {
    "SERVICE": "service",
    "MARKETING": "marketing",
    "PRIVACY": "privacy",
    "PRIVACY_CONSENT": "privacy_consent",
    "LOCATION": "location",
    "FEE": "fee",
}

def _build_s3_key(target: str, term_type: str, filename: str) -> str:
    """terms/{user|owner}/{category}/{filename}"""
    s3_target = "user" if target == "user" else "owner"
    category = _TERM_TYPE_TO_CATEGORY.get(term_type, term_type.lower())
    return f"terms/{s3_target}/{category}/{filename}"


def _parse_terms_filename(filename: str, target: str):
    """filename(예: service_term_260101.html) -> (term_type, version, effective_date, notice_date). 실패 시 None."""
    from datetime import datetime, timedelta
    if not filename or not filename.endswith(".html"):
        return None
    base = filename[:-5]
    parts = base.rsplit("_", 1)
    if len(parts) != 2:
        return None
    prefix, version_str = parts[0], parts[1]
    if len(version_str) != 6 or not version_str.isdigit():
        return None
    prefix_map = _FILENAME_PREFIX_TO_TERM_TYPE_USER if target == "user" else _FILENAME_PREFIX_TO_TERM_TYPE_PARTNER
    term_type = prefix_map.get(prefix)
    if not term_type:
        return None
    try:
        y = 2000 + int(version_str[:2])
        m = int(version_str[2:4])
        d = int(version_str[4:6])
        effective_date = datetime(y, m, d).date()
        notice_date = effective_date - timedelta(days=30)
    except (ValueError, TypeError):
        return None
    return {"term_type": term_type, "version": version_str, "effective_date": effective_date, "notice_date": notice_date}


@router.post("/terms/upload")
def upload_term_file(body: dict, user=Depends(verify_firebase_token)):
    """S3에 약관 txt 업로드 후 DB에 term/term_version 등록. Body: target(user|owner), filename, content [, title, notice_date, effective_date, reagreement_required]"""
    target = body.get("target")
    filename = body.get("filename")
    content = body.get("content", "")
    if not target or target not in ("user", "owner"):
        raise HTTPException(status_code=400, detail="target must be 'user' or 'owner'")
    if not filename:
        raise HTTPException(status_code=400, detail="filename is required")
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="invalid filename")

    parsed = _parse_terms_filename(filename, target)
    if not parsed:
        raise HTTPException(
            status_code=400,
            detail="filename must be like service_term_260101.html or partner_service_term_260101.html",
        )
    term_type = parsed["term_type"]
    version = parsed["version"]
    effective_date = parsed["effective_date"]
    notice_date = parsed["notice_date"]
    if body.get("effective_date"):
        try:
            from datetime import datetime
            effective_date = datetime.strptime(body["effective_date"], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass
    if body.get("notice_date"):
        try:
            from datetime import datetime
            notice_date = datetime.strptime(body["notice_date"], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass
    title = body.get("title") or _TERM_TYPE_DEFAULT_TITLE.get(term_type, term_type)
    reagreement_required = body.get("reagreement_required", True)
    if isinstance(reagreement_required, str):
        reagreement_required = reagreement_required.lower() in ("true", "1", "on", "yes")

    connection = get_db_connection()
    try:
        terms_list = terms_crud.get_all_terms_with_versions(connection, target)
        term_id = None
        for t in terms_list:
            if t.get("term_type") == term_type:
                term_id = t["id"]
                break
        if not term_id:
            term_id = terms_crud.create_term(connection, term_type, title, required=True, target=target)
        try:
            terms_crud.create_term_version(
                connection, term_id, version, notice_date, effective_date, reagreement_required
            )
        except ValueError as e:
            if "이미 존재하는 버전입니다" not in str(e):
                close_db_connection(connection)
                raise HTTPException(status_code=400, detail=str(e))
            # 버전이 이미 있으면 DB는 건너뛰고 S3만 덮어쓰기
    except ValueError as e:
        close_db_connection(connection)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        close_db_connection(connection)
        print(f"Error in terms/upload DB: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            close_db_connection(connection)
        except Exception:
            pass

    key = _build_s3_key(target, term_type, filename)
    try:
        body_bytes = content.encode("utf-8")
        S3_CLIENT.put_object(Bucket=TERMS_BUCKET_NAME, Key=key, Body=body_bytes, ContentType="text/html; charset=utf-8")
        return {"message": "Uploaded", "key": key, "term_type": term_type, "version": version}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/test/store")
def create_test_store(store: StoreCreate, user=Depends(verify_firebase_token)):
    """테스트 매장 추가"""
    try:
        store_id = store_crud.create_store(store)

        logo_key = f'store_logo/store_logo_{store_id}_{uuid.uuid4().hex[:8]}.png'
        bankbook_key = f'bankbook/bankbook_{store_id}_{uuid.uuid4().hex[:8]}.png'
        business_key = f'business_registration/business_registration_{store_id}_{uuid.uuid4().hex[:8]}.png'

        conn2 = get_db_connection()
        try:
            cur2 = conn2.cursor()
            cur2.execute(
                "UPDATE store SET store_logo_key = %s, bankbook_key = %s, business_registration_key = %s WHERE id = %s",
                (logo_key, bankbook_key, business_key, store_id)
            )
            conn2.commit()
        finally:
            close_db_connection(conn2)

        store_logo_url = S3_CLIENT.generate_presigned_url('put_object',
            Params={'Bucket': BUCKET_NAME, 'Key': logo_key}, ExpiresIn=3600)
        bankBook_put_url = S3_CLIENT.generate_presigned_url('put_object',
            Params={'Bucket': BUCKET_NAME, 'Key': bankbook_key}, ExpiresIn=3600)
        business_put_url = S3_CLIENT.generate_presigned_url('put_object',
            Params={'Bucket': BUCKET_NAME, 'Key': business_key}, ExpiresIn=3600)

        return {
            'store_id': store_id,
            'store_logo_url': store_logo_url,
            'bankBook_put_url': bankBook_put_url,
            'business_put_url': business_put_url
        }
    except Exception as e:
        print(f"Error in create_test_store: {traceback.format_exc()}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"failed register store: {str(e)}"
        )


@router.post("/test/menu/{store_id}")
def create_test_menu(store_id: int, menu: Menu, user=Depends(verify_firebase_token)):
    """테스트 메뉴 추가"""
    try:
        if menu.store_id != store_id:
            raise HTTPException(status_code=400, detail="store_id in path and body must match")
        
        menu_id = menu_crud.create_menu(store_id, menu)
        s3_urls = menu_crud.generate_menu_s3_urls(store_id, menu_id)
        
        return {
            'menu_id': menu_id,
            'menu_put_url': s3_urls['menu_put_url'],
            'menu_get_url': s3_urls['menu_get_url']
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in create_test_menu: {traceback.format_exc()}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"failed add menu: {str(e)}"
        )


@router.get("/promotions")
def get_all_promotions(active_only: bool = Query(False, description="활성 프로모션만 조회"), user=Depends(verify_firebase_token)):
    """전체 프로모션 리스트 조회 (적용 매장 수 포함)"""
    try:
        from crud import promotion as promotion_crud
        promotions = promotion_crud.get_all_fee_promotions(active_only=active_only)
        return {'promotions': promotions}
    except Exception as e:
        print(f"Error in get_all_promotions: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/promotions")
def create_fee_promotion(promotion: dict, user=Depends(verify_firebase_token)):
    """프로모션 생성

    Body:
        - title: 프로모션 제목
        - promo_type: 'FIXED_PERIOD' | 'PER_STORE_PERIOD' (기본 FIXED_PERIOD)
        - promo_fee_rate: 프로모션 수수료율 (%)
        - start_date: 시작일 (YYYY-MM-DD) — FIXED_PERIOD 필수
        - end_date: 종료일 (YYYY-MM-DD) — FIXED_PERIOD 필수
        - store_ids: 적용할 매장 ID 목록 (list[int], 선택, FIXED_PERIOD만 사용)
    """
    try:
        from crud import promotion as promotion_crud
        from datetime import datetime

        title = (promotion.get('title') or '').strip()
        promo_type = (promotion.get('promo_type') or promotion_crud.PROMO_TYPE_FIXED).strip()
        promo_fee_rate = promotion.get('promo_fee_rate')
        start_date_str = promotion.get('start_date')
        end_date_str = promotion.get('end_date')
        store_ids = promotion.get('store_ids') or []

        if not title or promo_fee_rate is None:
            raise HTTPException(status_code=400, detail="title, promo_fee_rate are required")

        if not isinstance(store_ids, list):
            raise HTTPException(status_code=400, detail="store_ids must be a list")

        start_date = None
        end_date = None
        if promo_type == promotion_crud.PROMO_TYPE_FIXED:
            if not start_date_str or not end_date_str:
                raise HTTPException(status_code=400, detail="FIXED_PERIOD는 start_date, end_date가 필수입니다.")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        promo_id = promotion_crud.create_fee_promotion(
            store_ids=store_ids,
            promo_fee_rate=float(promo_fee_rate),
            title=title,
            promo_type=promo_type,
            start_date=start_date,
            end_date=end_date,
        )
        return {'message': 'Promotion created successfully', 'promo_id': promo_id}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error in create_fee_promotion: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/promotions/{promo_id}")
def get_promotion_detail(promo_id: int, user=Depends(verify_firebase_token)):
    """프로모션 상세 조회 (적용 매장 목록 포함)"""
    try:
        from crud import promotion as promotion_crud
        detail = promotion_crud.get_fee_promotion_detail(promo_id)
        if not detail:
            raise HTTPException(status_code=404, detail="Promotion not found")
        return detail
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_promotion_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/promotions/{promo_id}")
def delete_fee_promotion(promo_id: int, user=Depends(verify_firebase_token)):
    """프로모션 삭제"""
    try:
        from crud import promotion as promotion_crud
        deleted = promotion_crud.delete_fee_promotion(promo_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Promotion not found")
        return {'message': 'Promotion deleted successfully'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in delete_fee_promotion: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stores/{store_id}/promotions")
def get_store_promotions(store_id: int, user=Depends(verify_firebase_token)):
    """매장별 프로모션 리스트 조회 (활성 + 이력 통합, 최신순)"""
    try:
        from crud import promotion as promotion_crud
        promotions = promotion_crud.get_promotions_by_store(store_id)
        return {'promotions': promotions}
    except Exception as e:
        print(f"Error in get_store_promotions: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/promotions/{store_id}/history")
def get_store_promotion_history(
    store_id: int,
    page: int = Query(1, ge=1),
    limit: int = Query(5, ge=1, le=100),
    user=Depends(verify_firebase_token),
):
    """매장 프로모션 이력 조회 (페이지네이션)"""
    try:
        from crud import promotion as promotion_crud
        result = promotion_crud.get_fee_promotions_by_store(store_id, page=page, limit=limit)
        return result
    except Exception as e:
        print(f"Error in get_store_promotion_history: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/stores/{store_id}/promotions/{promo_id}/apply")
def apply_promotion_to_store(store_id: int, promo_id: int, body: dict = None, user=Depends(verify_firebase_token)):
    """매장에 프로모션 등록

    Body (PER_STORE_PERIOD만 사용):
        - start_date: 시작일 (YYYY-MM-DD)
        - end_date: 종료일 (YYYY-MM-DD)
    """
    try:
        from crud import promotion as promotion_crud
        from datetime import datetime

        body = body or {}
        start_date_str = body.get('start_date')
        end_date_str = body.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

        promotion_crud.apply_promotion_to_store(promo_id, store_id, start_date=start_date, end_date=end_date)
        return {'message': 'Promotion applied successfully'}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error in apply_promotion_to_store: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/stores/{store_id}/promotions/{promo_id}")
def remove_promotion_from_store(store_id: int, promo_id: int, user=Depends(verify_firebase_token)):
    """매장 프로모션 등록 해제 (soft delete)"""
    try:
        from crud import promotion as promotion_crud
        removed = promotion_crud.remove_promotion_from_store(promo_id, store_id)
        if not removed:
            raise HTTPException(status_code=404, detail="활성 상태인 프로모션 매핑을 찾을 수 없습니다.")
        return {'message': 'Promotion removed successfully'}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in remove_promotion_from_store: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/statistics/gifticons")
def get_admin_gifticon_statistics(user=Depends(verify_firebase_token)):
    """관리자 통계 데이터 조회 (전체 발행 수, 사용 수, 미사용 수)"""
    connection = get_db_connection()
    try:
        from crud import stats as stats_crud
        result = stats_crud.get_admin_statistics()
        return result
    except Exception as e:
        print(f"Error in get_admin_gifticon_statistics: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/statistics/settlement")
def get_admin_settlement_statistics(
    start_date: Optional[str] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="종료일 (YYYY-MM-DD)"),
    user=Depends(verify_firebase_token),
):
    """관리자 정산 데이터 조회 (정산금액, 플랫폼 수수료 매출)"""
    connection = get_db_connection()
    try:
        from crud import stats as stats_crud
        from datetime import datetime
        
        start = None
        end = None
        
        if start_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        result = stats_crud.get_admin_settlement_data(start, end)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        print(f"Error in get_admin_settlement_statistics: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/settlement/cycles")
def get_settlement_cycles(
    status: Optional[str] = Query(None, description="'OPEN' 또는 'CLOSED', None이면 전체"),
    start_date: Optional[str] = Query(None, description="조회 시작일 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="조회 종료일 (YYYY-MM-DD)"),
    user=Depends(verify_firebase_token),
):
    """정산 주기 리스트 조회"""
    from datetime import date as date_type
    connection = get_db_connection()
    try:
        parsed_start = None
        parsed_end = None
        if start_date:
            try:
                parsed_start = date_type.fromisoformat(start_date)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Invalid start_date format: {start_date}")
        if end_date:
            try:
                parsed_end = date_type.fromisoformat(end_date)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Invalid end_date format: {end_date}")
        from crud import settlement_cycle as cycle_crud
        cycles = cycle_crud.get_settlement_cycles(status, parsed_start, parsed_end)
        return {'cycles': cycles}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_settlement_cycles: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/settlement/cycles/{cycle_id}")
def get_settlement_cycle(cycle_id: int, user=Depends(verify_firebase_token)):
    """정산 주기 상세 조회"""
    connection = get_db_connection()
    try:
        from crud import settlement_cycle as cycle_crud
        cycle = cycle_crud.get_settlement_cycle_by_id(cycle_id)
        if not cycle:
            raise HTTPException(status_code=404, detail="Settlement cycle not found")
        return cycle
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_settlement_cycle: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


def _close_settlement_cycle_impl(cycle_id: int):
    """정산 주기 마감 구현 (PATCH/POST 공용)"""
    from crud import settlement_cycle as cycle_crud
    updated = cycle_crud.close_settlement_cycle(cycle_id)
    if not updated:
        raise HTTPException(status_code=404, detail="정산 주기를 찾을 수 없거나 이미 마감되었습니다.")
    return {"success": True, "message": "정산 주기가 마감되었습니다.", "cycle_id": cycle_id}


@router.patch("/settlement/cycles/{cycle_id}/close")
def close_settlement_cycle_patch(cycle_id: int, user=Depends(verify_firebase_token)):
    """정산 주기 마감: settlement_cycles.status 를 CLOSED 로 변경 (PATCH)"""
    try:
        return _close_settlement_cycle_impl(cycle_id)
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in close_settlement_cycle: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/settlement/cycles/{cycle_id}/close")
def close_settlement_cycle_post(cycle_id: int, user=Depends(verify_firebase_token)):
    """정산 주기 마감 (POST - 프록시에서 PATCH 미지원 시 사용)"""
    try:
        return _close_settlement_cycle_impl(cycle_id)
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in close_settlement_cycle: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/settlement/cycles/{cycle_id}/status")
@router.post("/settlement/cycles/{cycle_id}/status")
def update_settlement_cycle_status(cycle_id: int, status: str = Query(..., description="변경할 상태: 'OPEN' 또는 'CLOSED'"), user=Depends(verify_firebase_token)):
    """정산 주기 상태 변경 (OPEN ↔ CLOSED)"""
    try:
        from crud import settlement_cycle as cycle_crud
        new_status = cycle_crud.update_settlement_cycle_status(cycle_id, status)
        if new_status is None:
            raise HTTPException(status_code=404, detail="정산 주기를 찾을 수 없습니다.")
        return {"success": True, "cycle_id": cycle_id, "status": new_status}
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error in update_settlement_cycle_status: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/settlement/cycles/generate")
def generate_settlement_cycles(
    start_date: Optional[str] = Query(None, description="시작일 (YYYY-MM-DD), 기본값: 오늘"),
    months: int = Query(12, ge=1, le=24, description="생성할 개월 수 (1-24)"),
    user=Depends(verify_firebase_token),
):
    """정산 주기 데이터 생성 (1년치)"""
    connection = get_db_connection()
    try:
        from crud import settlement_cycle as cycle_crud
        from datetime import date, datetime
        
        if start_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start = date.today()
        
        count = cycle_crud.generate_settlement_cycles(start, months)
        return {
            'message': f'{count}개의 정산 주기가 생성되었습니다.',
            'count': count,
            'start_date': start.isoformat()
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        print(f"Error in generate_settlement_cycles: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/settlement/list")
def get_settlement_list_by_cycle(
    cycle_id: int = Query(..., description="정산 주기 ID"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(10, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """정산 주기별 매장 정산 리스트 (페이지네이션)"""
    try:
        from crud import settlement as settlement_crud
        return settlement_crud.get_settlements_by_cycle(cycle_id, page, limit)
    except Exception as e:
        print(f"Error in get_settlement_list_by_cycle: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/settlement/{settlement_id}/details")
def get_settlement_details_admin(
    settlement_id: int,
    detail_page: int = Query(1, ge=1, description="건별 내역 페이지 번호"),
    detail_limit: int = Query(10, ge=1, le=100, description="건별 내역 페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """정산 상세 (헤더 + 건별 내역, 페이지네이션)"""
    try:
        from crud import settlement as settlement_crud
        data = settlement_crud.get_settlement_detail_for_admin(settlement_id, detail_page, detail_limit)
        if not data:
            raise HTTPException(status_code=404, detail="Settlement not found")
        return data
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_settlement_details_admin: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/settlement/{settlement_id}/status")
@router.post("/settlement/{settlement_id}/status")
def update_settlement_status(settlement_id: int, body: dict, user=Depends(verify_firebase_token)):
    """정산 상태 변경. body: { "status": "PAID", "failure_reason": "..." }"""
    status = (body.get("status") or "").strip().upper()
    if not status:
        raise HTTPException(status_code=400, detail="status is required")
    failure_reason = body.get("failure_reason")
    try:
        from crud import settlement as settlement_crud
        updated = settlement_crud.update_settlement_status(settlement_id, status, failure_reason)
        if not updated:
            raise HTTPException(status_code=404, detail="Settlement not found")

        if status == "COMPLETED":
            try:
                from app.aligo_service import send_settlement_complete
                import pymysql
                from db.session import get_db_connection, close_db_connection
                conn = get_db_connection()
                cur = conn.cursor(pymysql.cursors.DictCursor)
                try:
                    cur.execute("""
                        SELECT s.store_id, st.store_name,
                               s.period_start, s.period_end,
                               s.net_payout_amount,
                               COALESCE(s.bank_name, a.bank) AS bank_name,
                               COALESCE(s.account_number, a.account) AS account_number,
                               o.phone
                        FROM settlement s
                        JOIN store st ON s.store_id = st.id
                        JOIN owner o ON st.owner_id = o.id
                        LEFT JOIN account a ON s.store_id = a.store_id
                        WHERE s.settlement_id = %s
                    """, (settlement_id,))
                    row = cur.fetchone()
                finally:
                    cur.close()
                    close_db_connection(conn)

                if row and row.get("phone"):
                    period = f"{row['period_start'].strftime('%Y.%m.%d')} ~ {row['period_end'].strftime('%Y.%m.%d')}"
                    amount = f"{int(row['net_payout_amount'] or 0):,}"
                    send_settlement_complete(
                        receiver=row["phone"],
                        store_name=row["store_name"] or "",
                        period=period,
                        amount=amount,
                        bank_name=row["bank_name"] or "",
                        account_number=row["account_number"] or "",
                    )
            except Exception:
                print(f"[알림톡] 정산 완료 발송 실패: {traceback.format_exc()}")

        return {"success": True, "settlement_id": settlement_id, "status": status}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_settlement_status: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/settlement/{settlement_id}/tax-invoice")
@router.post("/settlement/{settlement_id}/tax-invoice")
def update_settlement_tax_invoice(settlement_id: int, body: dict, user=Depends(verify_firebase_token)):
    """세금계산서 발행 여부 변경. body: { "tax_invoice_issued": true }"""
    if "tax_invoice_issued" not in body:
        raise HTTPException(status_code=400, detail="tax_invoice_issued is required")
    tax_invoice_issued = bool(body["tax_invoice_issued"])
    try:
        from crud import settlement as settlement_crud
        updated = settlement_crud.update_settlement_tax_invoice(settlement_id, tax_invoice_issued)
        if not updated:
            raise HTTPException(status_code=404, detail="Settlement not found")
        return {"success": True, "settlement_id": settlement_id, "tax_invoice_issued": tax_invoice_issued}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_settlement_tax_invoice: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/settlement/create/{cycle_id}")
def create_settlement_data(cycle_id: int, user=Depends(verify_firebase_token)):
    """정산 데이터 생성 (정산 주기별)
    
    cycle_id는 /admin/settlement/cycles API로 조회 가능합니다.
    """
    connection = get_db_connection()
    try:
        from crud import stats as stats_crud
        result = stats_crud.create_settlement_data(cycle_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error in create_settlement_data: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/refund/list")
def get_refund_list_api(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    refund_type: Optional[str] = Query(None, description="PURCHASER 또는 RECEIVER"),
    user=Depends(verify_firebase_token),
):
    """환불 리스트 (관리자). id, 구매날짜, 환불요청날짜, 환불타입, 예금주, 계좌번호, 지급상태"""
    try:
        from crud import refund as refund_crud
        result = refund_crud.get_refund_list(page=page, limit=limit, refund_type=refund_type)
        return result
    except Exception as e:
        print(f"Error in get_refund_list: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/refund/{refund_id}/status")
def update_refund_status_api(refund_id: int, body: dict, user=Depends(verify_firebase_token)):
    """환불 지급상태 변경. body: { \"status\": \"COMPLETED\" } (REQUESTED, COMPLETED, FAILED)"""
    status = (body.get("status") or "").strip().upper()
    if not status:
        raise HTTPException(status_code=400, detail="status is required")
    try:
        from crud import refund as refund_crud
        updated = refund_crud.update_refund_status(refund_id, status)
        if not updated:
            raise HTTPException(status_code=404, detail="Refund not found or invalid status")
        return updated
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_refund_status: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/owners")
def get_owners(
    search: Optional[str] = Query(None, description="이름, 이메일, 전화번호, ID로 검색"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지당 항목 수"),
    user=Depends(verify_firebase_token),
):
    """사장님 리스트 (관리자용, 페이지네이션)"""
    connection = get_db_connection()
    try:
        result = admin_crud.get_owners(connection, search, page, limit)
        return {
            'owners': result['items'],
            'pagination': {
                'total': result['total'],
                'page': result['page'],
                'limit': result['limit'],
                'total_pages': result['total_pages'],
            },
        }
    except Exception as e:
        print(f"Error in get_owners: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/owners/{owner_id}")
def get_owner_detail(owner_id: int, user=Depends(verify_firebase_token)):
    """사장님 상세 정보"""
    connection = get_db_connection()
    try:
        owner = admin_crud.get_owner_detail(connection, owner_id)
        if not owner:
            raise HTTPException(status_code=404, detail="Owner not found")
        return owner
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_owner_detail: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


# ── 앱 버전 관리 ──────────────────────────────────────────────────────────────

from pydantic import BaseModel as _BaseModel
from typing import Optional as _Optional, Literal as _Literal
import pymysql as _pymysql

class AppVersionCreate(_BaseModel):
    platform: _Literal['ios', 'android']
    version: str
    memo: _Optional[str] = None


@router.get("/app-versions")
def list_app_versions(platform: _Optional[str] = None, user=Depends(verify_firebase_token)):
    """앱 버전 목록 조회 (매니저용)"""
    connection = get_db_connection()
    try:
        cursor = connection.cursor(_pymysql.cursors.DictCursor)
        if platform:
            cursor.execute(
                "SELECT * FROM app_versions WHERE platform = %s ORDER BY created_at DESC",
                (platform,)
            )
        else:
            cursor.execute("SELECT * FROM app_versions ORDER BY created_at DESC")
        rows = cursor.fetchall()
        for r in rows:
            r['is_force_update'] = bool(r['is_force_update'])
            if r.get('created_at'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return rows
    except Exception as e:
        print(f"Error in list_app_versions: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.post("/app-versions", status_code=201)
def create_app_version(body: AppVersionCreate, user=Depends(verify_firebase_token)):
    """앱 버전 등록"""
    connection = get_db_connection()
    try:
        cursor = connection.cursor(_pymysql.cursors.DictCursor)
        cursor.execute(
            "INSERT INTO app_versions (platform, version, is_force_update, memo) VALUES (%s, %s, 0, %s)",
            (body.platform, body.version, body.memo)
        )
        connection.commit()
        new_id = cursor.lastrowid
        cursor.execute("SELECT * FROM app_versions WHERE id = %s", (new_id,))
        row = cursor.fetchone()
        row['is_force_update'] = bool(row['is_force_update'])
        if row.get('created_at'):
            row['created_at'] = row['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return row
    except Exception as e:
        print(f"Error in create_app_version: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.patch("/app-versions/{version_id}")
def update_app_version_force(version_id: int, is_force_update: bool, user=Depends(verify_firebase_token)):
    """강제업데이트 여부 변경"""
    connection = get_db_connection()
    try:
        cursor = connection.cursor(_pymysql.cursors.DictCursor)
        cursor.execute(
            "UPDATE app_versions SET is_force_update = %s WHERE id = %s",
            (1 if is_force_update else 0, version_id)
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Version not found")
        connection.commit()
        cursor.execute("SELECT * FROM app_versions WHERE id = %s", (version_id,))
        row = cursor.fetchone()
        row['is_force_update'] = bool(row['is_force_update'])
        if row.get('created_at'):
            row['created_at'] = row['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return row
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_app_version_force: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


# ── Popup Admin API ────────────────────────────────────────────────────────────
from models.popup import PopupCreate, PopupUpdate


@router.post("/popups/image")
def popup_image_upload(target_type: str, user=Depends(verify_firebase_token)):
    """팝업 이미지 S3 presigned URL 발급 (target_type: user | owner)"""
    if target_type not in ('user', 'owner'):
        raise HTTPException(status_code=400, detail="target_type must be 'user' or 'owner'")
    key = f"{target_type}/popup/{uuid.uuid4().hex}.jpg"
    put_url = S3_CLIENT.generate_presigned_url(
        'put_object',
        Params={'Bucket': BUCKET_NAME, 'Key': key, 'ContentType': 'image/jpeg'},
        ExpiresIn=3600
    )
    image_url = f"https://{BUCKET_NAME}.s3.ap-northeast-2.amazonaws.com/{key}"
    return {"put_url": put_url, "image_url": image_url, "key": key}


@router.get("/popups")
def list_popups(
    target_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user=Depends(verify_firebase_token)
):
    """팝업 목록 조회"""
    if target_type and target_type not in ('user', 'owner'):
        raise HTTPException(status_code=400, detail="target_type must be 'user' or 'owner'")
    connection = get_db_connection()
    try:
        return admin_crud.get_popups(connection, target_type, page, limit)
    except Exception as e:
        print(f"Error in list_popups: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.post("/popups", status_code=201)
def create_popup(body: PopupCreate, user=Depends(verify_firebase_token)):
    """팝업 생성 (display_order는 해당 target_type 마지막 순서 + 1 자동 부여)"""
    if body.target_type not in ('user', 'owner'):
        raise HTTPException(status_code=400, detail="target_type must be 'user' or 'owner'")
    connection = get_db_connection()
    try:
        import pymysql as _pym
        cursor = connection.cursor(_pym.cursors.DictCursor)
        cursor.execute(
            "SELECT COALESCE(MAX(display_order), -1) + 1 as next_order FROM popup WHERE target_type = %s",
            (body.target_type,)
        )
        next_order = cursor.fetchone()['next_order']
        cursor.close()
        return admin_crud.create_popup(
            connection,
            target_type=body.target_type,
            title=body.title,
            image_url=body.image_url,
            link_url=body.link_url,
            display_order=next_order,
            is_active=body.is_active,
            start_at=body.start_at,
            end_at=body.end_at,
        )
    except Exception as e:
        print(f"Error in create_popup: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.patch("/popups/reorder")
def reorder_popups(ordered_ids: list[int], user=Depends(verify_firebase_token)):
    """팝업 순서 일괄 변경 (id 배열 순서대로 display_order 0부터 재부여)"""
    if not ordered_ids:
        raise HTTPException(status_code=400, detail="ordered_ids is required")
    connection = get_db_connection()
    try:
        admin_crud.reorder_popups(connection, ordered_ids)
        return {"message": "순서가 업데이트되었습니다."}
    except Exception as e:
        print(f"Error in reorder_popups: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.get("/popups/{popup_id}")
def get_popup(popup_id: int, user=Depends(verify_firebase_token)):
    """팝업 상세 조회"""
    connection = get_db_connection()
    try:
        row = admin_crud.get_popup(connection, popup_id)
        if not row:
            raise HTTPException(status_code=404, detail="Popup not found")
        return row
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_popup: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.put("/popups/{popup_id}")
def update_popup(popup_id: int, body: PopupUpdate, user=Depends(verify_firebase_token)):
    """팝업 수정"""
    connection = get_db_connection()
    try:
        row = admin_crud.update_popup(connection, popup_id, **body.model_dump(exclude_unset=True))
        if not row:
            raise HTTPException(status_code=404, detail="Popup not found")
        return row
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_popup: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.delete("/popups/{popup_id}", status_code=204)
def delete_popup(popup_id: int, user=Depends(verify_firebase_token)):
    """팝업 삭제"""
    connection = get_db_connection()
    try:
        deleted = admin_crud.delete_popup(connection, popup_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Popup not found")
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in delete_popup: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)


@router.patch("/popups/{popup_id}/toggle")
def toggle_popup(popup_id: int, user=Depends(verify_firebase_token)):
    """팝업 활성화/중지 토글"""
    connection = get_db_connection()
    try:
        row = admin_crud.toggle_popup(connection, popup_id)
        if not row:
            raise HTTPException(status_code=404, detail="Popup not found")
        return row
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in toggle_popup: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        close_db_connection(connection)
